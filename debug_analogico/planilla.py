#!/usr/bin/env python3
"""Escribe la planilla de medicion como Excel, con las verificaciones hechas.

La gracia sobre el CSV es que la columna `ok` es una formula: se completa
sola al escribir la medicion, comparandola contra el rango de tolerancia.
Las celdas de entrada quedan resaltadas.

Ademas de la lectura final se anotan la lectura inicial y cuanto se espero.
Eso permite reconstruir despues por que una medida dio raro: si la inicial
coincide con la columna `arranca kohm` y se espero poco, la lectura quedo a
mitad de camino del transitorio de C1 y no hay nada roto.

Se usa desde red_analogica.py; no se ejecuta solo.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_CABECERA = PatternFill("solid", fgColor="1F3864")
_ENTRADA = PatternFill("solid", fgColor="FFF2CC")
_VERDE = PatternFill("solid", fgColor="C6EFCE")
_ROJO = PatternFill("solid", fgColor="FFC7CE")
_AMBAR = PatternFill("solid", fgColor="FFEB9C")
_BORDE = Border(bottom=Side(style="thin", color="BFBFBF"))

# Indices (base 0) de las celdas que completa el usuario, por hoja.
_ANOTADAS_R = {"medido_inicial": 7, "espera_s": 8, "medido": 9,
               "esperado_corregido": 11, "notas_medicion": 12}
_ANOTADAS_C = {"medido_inicial": 7, "espera_s": 8, "medido": 9,
               "notas_medicion": 11}


def _unidad_c(faradios):
    """Unidad comoda para cada capacitor, para no escribir 0.000177."""
    if faradios >= 1e-6:
        return "uF", 1e-6
    if faradios >= 1e-9:
        return "nF", 1e-9
    return "pF", 1e-12


def _encabezar(ws, titulos, anchos):
    ws.append(titulos)
    for i, (t, a) in enumerate(zip(titulos, anchos), start=1):
        c = ws.cell(row=1, column=i)
        c.fill = _CABECERA
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32


def _pintar_ok(ws, col_ok, ultima):
    """Verde si entra en tolerancia, rojo si no, ambar si falta esperar."""
    rango = "%s2:%s%d" % (col_ok, col_ok, ultima)
    ws.conditional_formatting.add(rango, CellIsRule(
        operator="equal", formula=['"OK"'], fill=_VERDE))
    ws.conditional_formatting.add(rango, CellIsRule(
        operator="equal", formula=['"ESPERAR"'], fill=_AMBAR))
    for mal in ('"FUERA"', '"CORTO?"'):
        ws.conditional_formatting.add(rango, CellIsRule(
            operator="equal", formula=[mal], fill=_ROJO))


def _marcar_entrada(ws, fila, columnas, ancho_total, col_ok):
    for col in columnas:
        ws.cell(fila, col).fill = _ENTRADA
    ws.cell(fila, col_ok).font = Font(bold=True)
    ws.cell(fila, col_ok).alignment = Alignment(horizontal="center")
    for col in range(1, ancho_total + 1):
        ws.cell(fila, col).border = _BORDE


def _hoja_resistencia(wb, filas, previo, es_base):
    ws = wb.create_sheet("Resistencia")
    _encabezar(ws,
               ["base", "par", "esperado kohm", "min kohm", "max kohm",
                "camino", "arranca kohm", "1a lectura kohm", "espera s",
                "medido kohm", "ok", "esperado corregido", "notas"],
               [6, 34, 12, 10, 10, 22, 11, 12, 9, 12, 10, 14, 30])

    for f in filas:
        r = ws.max_row + 1
        abierto = f["esperado"] == "abierto"
        guardado = previo.get(("R", f["par"]), {})

        ws.cell(r, 1, "*" if es_base(f["senal_a"], f["senal_b"]) else "")
        ws.cell(r, 2, f["par"])
        ws.cell(r, 3, "abierto" if abierto
                else round(float(f["esperado_ohm"]) / 1e3, 3))
        if not abierto:
            ws.cell(r, 4, round(float(f["min_ohm"]) / 1e3, 3))
            ws.cell(r, 5, round(float(f["max_ohm"]) / 1e3, 3))
        ws.cell(r, 6, f["camino"])
        if f["lectura_inicial"]:
            ws.cell(r, 7, float(f["lectura_inicial"].replace("k", "").strip()))
        ws.cell(r, 8, guardado.get("medido_inicial", ""))
        ws.cell(r, 9, guardado.get("espera_s", ""))
        ws.cell(r, 10, guardado.get("medido", ""))

        # Con rango: dentro o fuera. Sin rango (abierto): un numero es corto.
        # Si la fila tiene transitorio y se espero poco, avisa antes de
        # declarar la medida mala.
        ws.cell(r, 11, (
            '=IF(J{r}="","",'
            'IF(D{r}="",IF(ISNUMBER(J{r}),"CORTO?","OK"),'
            'IF(AND(J{r}>=D{r},J{r}<=E{r}),"OK",'
            'IF(AND(G{r}<>"",I{r}<>"",I{r}<60),"ESPERAR","FUERA"))))'
        ).format(r=r))
        ws.cell(r, 12, guardado.get("esperado_corregido", ""))
        ws.cell(r, 13, guardado.get("notas_medicion", "") or f["nota"])

        _marcar_entrada(ws, r, (8, 9, 10, 12, 13), 13, 11)

    _pintar_ok(ws, "K", ws.max_row)
    ws.auto_filter.ref = "A1:M%d" % ws.max_row
    return ws


def _hoja_capacitancia(wb, caps, previo, faradios):
    ws = wb.create_sheet("Capacitancia")
    _encabezar(ws,
               ["par", "componentes", "unidad", "esperado", "min", "max",
                "en paralelo", "1a lectura", "espera s", "medido", "ok",
                "notas"],
               [34, 12, 8, 10, 10, 10, 26, 11, 9, 11, 10, 30])

    for c in caps:
        r = ws.max_row + 1
        total = sum(faradios(v) for v in c["valores"])
        unidad, escala = _unidad_c(total)
        guardado = previo.get(("C", c["par"]), {})

        ws.cell(r, 1, c["par"])
        ws.cell(r, 2, c["componentes"])
        ws.cell(r, 3, unidad)
        ws.cell(r, 4, round(total / escala, 4))
        ws.cell(r, 5, round(total * 0.8 / escala, 4))
        ws.cell(r, 6, round(total * 1.2 / escala, 4))
        ws.cell(r, 7, "nada: lectura limpia" if c["shunt"] is None
                else "%.1f kohm: lectura dudosa" % (c["shunt"] / 1e3))
        ws.cell(r, 8, guardado.get("medido_inicial", ""))
        ws.cell(r, 9, guardado.get("espera_s", ""))
        ws.cell(r, 10, guardado.get("medido", ""))
        ws.cell(r, 11, ('=IF(J{r}="","",IF(AND(J{r}>=E{r},J{r}<=F{r}),'
                        '"OK","FUERA"))').format(r=r))
        ws.cell(r, 12, guardado.get("notas_medicion", ""))

        _marcar_entrada(ws, r, (8, 9, 10, 12), 12, 11)

    _pintar_ok(ws, "K", ws.max_row)
    return ws


def _hoja_instrucciones(wb, rv1_pct):
    ws = wb.create_sheet("Antes de medir", 0)
    ws.column_dimensions["A"].width = 100
    lineas = [
        ("Como usar esta planilla", True),
        ("", False),
        ("Solo se completan las celdas amarillas. La columna 'ok' es una "
         "formula: se pone verde", False),
        ("si la medicion entra en tolerancia y roja si no. No hay que "
         "comparar a mano.", False),
        ("Las resistencias van en kohm y los capacitores en la unidad que "
         "indica su fila.", False),
        ("En la hoja Resistencia, un * marca los doce chequeos base: si "
         "queres hacer solo esos,", False),
        ("segui los asteriscos.", False),
        ("", False),
        ("Por que se anota la primera lectura y la espera", True),
        ("", False),
        ("El electrolitico C1 de 680 uF da un camino en paralelo mientras se "
         "carga con la", False),
        ("corriente del ohmetro, y se abre a medida que se carga. La columna "
         "'arranca kohm'", False),
        ("dice desde donde deberia arrancar cada lectura afectada; sube al "
         "valor final en", False),
        ("uno o dos minutos (tau ~ 17 s).", False),
        ("", False),
        ("Anotando la '1a lectura' y la 'espera s' queda registrado el "
         "transitorio completo,", False),
        ("asi que despues se puede reconstruir una medida rara sin volver a "
         "la mesa: si la", False),
        ("primera lectura coincide con 'arranca kohm', el circuito esta bien "
         "y lo que fallo", False),
        ("fue la paciencia. Si NO coincide, ahi si hay algo distinto de lo "
         "que dice el modelo.", False),
        ("", False),
        ("Por eso, si una medida cae fuera de rango pero se espero menos de "
         "60 s en una fila", False),
        ("con transitorio, la columna 'ok' dice ESPERAR (ambar) en vez de "
         "FUERA (rojo).", False),
        ("", False),
        ("Las otras dos trampas", True),
        ("", False),
        ("1. El geofono queda en paralelo con los 100 kohm de entrada. "
         "Desconecta J4 o vas a", False),
        ("   leer la resistencia de bobina.", False),
        ("2. El trimmer RV1 esta al %.1f%% nominal, pero el valor real es el "
         "que tenga puesto" % rv1_pct, False),
        ("   el cursor. Medi BPo-SUMm, restale R7 y corregi RV1_FRACCION en "
         "el script.", False),
        ("", False),
        ("Lo que esta planilla no cubre", True),
        ("", False),
        ("C1 (680 uF) y C4 (47 nF) tienen una pata en un nodo interno: no "
         "hay dos pines entre", False),
        ("los cuales medirlos. Tampoco R4 (43 kohm), que C1 bloquea en "
         "continua; hay que", False),
        ("medirlo en sus propios pads. Eso lo cierra un barrido en "
         "frecuencia.", False),
        ("", False),
        ("Regenerar la planilla NO borra lo anotado: el script relee el "
         "archivo y arrastra", False),
        ("las celdas amarillas.", False),
    ]
    for i, (texto, titulo) in enumerate(lineas, start=1):
        c = ws.cell(i, 1, texto)
        if titulo:
            c.font = Font(bold=True, size=12, color="1F3864")
    return ws


def _rescatar(fila, indices, clave_par):
    par = fila[clave_par]
    if not par:
        return None, None
    datos = {}
    for nombre, idx in indices.items():
        valor = fila[idx] if idx < len(fila) else None
        datos[nombre] = "" if valor is None else valor
    return (par, datos) if any(datos.values()) else (None, None)


def leer_anotaciones(ruta):
    """Rescata lo ya escrito a mano para no pisarlo al regenerar."""
    try:
        wb = load_workbook(ruta, data_only=False)
    except (FileNotFoundError, OSError, KeyError):
        return {}

    previo = {}
    for hoja, tipo, indices, col_par in (
            ("Resistencia", "R", _ANOTADAS_R, 1),
            ("Capacitancia", "C", _ANOTADAS_C, 0)):
        if hoja not in wb.sheetnames:
            continue
        for fila in wb[hoja].iter_rows(min_row=2, values_only=True):
            par, datos = _rescatar(fila, indices, col_par)
            if par:
                previo[(tipo, par)] = datos
    return previo


def escribir(ruta, filas_r, caps, es_base, faradios, rv1_pct):
    """Genera el .xlsx completo. Devuelve cuantas filas traian anotacion."""
    previo = leer_anotaciones(ruta)
    wb = Workbook()
    wb.remove(wb.active)
    _hoja_instrucciones(wb, rv1_pct)
    _hoja_resistencia(wb, filas_r, previo, es_base)
    _hoja_capacitancia(wb, caps, previo, faradios)
    wb.active = 1
    wb.save(ruta)
    return len(previo)
